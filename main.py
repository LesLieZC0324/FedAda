"""
Client <--> Edge Server <--> Cloud Server
"""

from options import args_parser
from client import Client
from edge import Edge
from cloud import Cloud
from datasets.get_data import get_dataloaders
from models.mnistLR import mnistLR
from models.mnistCNN import mnistCNN
from models.LeNet import LeNet
from models.cifarCNN import cifarCNN
from models.cifarResNet import ResNet18, ResNet34, ResNet50, ResNet101, ResNet152
from models.mobileNet import MobileNet
from models.AlexNet import AlexNet
from models.ResNet9 import ResNet9

import openpyxl
import torch
from torch import nn
import copy
import numpy as np
from tqdm import tqdm
import os
import pandas as pd
import torch.backends.cudnn as cudnn
cudnn.banchmark = True
cudnn.enabled = True


def initialize_global_model(args):
    if args.dataset == 'mnist' or args.dataset == 'fmnist':
        if args.model == 'mnistLR' or args.model == 'fmnistLR':
            global_model = mnistLR(input_channels=1, output_channels=10)
        elif args.model == 'mnistCNN' or args.model == 'fmnistCNN':
            global_model = mnistCNN(input_channels=1, output_channels=10)
        elif args.model == 'MobileNet':
            global_model = MobileNet(input_channels=1, output_channels=10)
        elif args.model == 'LeNet':
            global_model = LeNet(input_channels=1, output_channels=10)
        else:
            raise ValueError(f"Model{args.model} not implemented for mnist/fmnist")
    elif args.dataset == 'cifar10' or args.dataset == 'cifar':
        if args.model == 'cifarCNN':
            global_model = cifarCNN(input_channels=3, output_channels=10)
        elif args.model == 'AlexNet':
            global_model = AlexNet(input_channels=3, output_channels=10)
        elif args.model == 'ResNet9':
            global_model = ResNet9(input_channels=3, output_channels=10)
        elif args.model == 'ResNet18':
            global_model = ResNet18(input_channels=3, output_channels=10)
        else:
            raise ValueError(f"Model{args.model} not implemented for cifar10")
    elif args.dataset == 'cifar100':
        if args.model == 'ResNet9':
            global_model = ResNet9(input_channels=3, output_channels=100)
        elif args.model == 'ResNet18':
            global_model = ResNet18(input_channels=3, output_channels=100)
        else:
            raise ValueError(f"Model{args.model} not implemented for cifar")
    else:
        raise ValueError('Wrong input for dataset, only mnist, fmnist and cifar10 are valid')
    total = sum([param.nelement() for param in global_model.parameters()]) / 1e6
    print("Number of parameter: {:.2f}M".format(total))
    return global_model, total


def all_client_test(server, clients, client_ids, device):
    for client_id in client_ids:
        server.send_to_client(clients[client_id])
        clients[client_id].sync_with_edge()

    correct_client = 0.0
    total_client = 0.0
    for client_id in client_ids:
        correct, total = clients[client_id].test_model(device)
        correct_client += correct
        total_client += total
    return correct_client, total_client


def global_model_test(v_test_loader, global_model, device):
    correct_all = 0.0
    total_all = 0.0
    with torch.no_grad():
        for data in v_test_loader:
            inputs, labels = data
            inputs = inputs.to(device)
            labels = labels.to(device)
            outputs = global_model(inputs)
            _, preds = torch.max(outputs, 1)
            correct_all += (preds == labels).sum().item()
            total_all += labels.size(0)
    return correct_all, total_all


def myFedAvg(args):
    torch.manual_seed(args.seed)
    np.random.seed(args.seed)
    if args.cuda:
        torch.cuda.manual_seed(args.seed)
        cuda_to_use = torch.device(f'cuda:{args.gpu}')
    device = cuda_to_use if torch.cuda.is_available() else "cpu"
    print(f'Using device {device}')

    # get train_dataset(need to split) and test_dataset
    train_loaders, test_loaders, v_train_loader, v_test_loader = get_dataloaders(args)

    # global model initialization
    global_model, total = initialize_global_model(args)
    if args.cuda:
        global_model = global_model.cuda(device)

    # client model initialization
    clients = []
    for i in range(args.num_clients):
        clients.append(Client(id=i, train_loader=train_loaders[i], test_loader=test_loaders[i],
                              device=device, global_model=global_model, args=args))
    initialization_parameters = list(clients[0].model.shared_layers.parameters())
    parameters_length = len(initialization_parameters)
    for client in clients:
        client_parameters = list(client.model.shared_layers.parameters())
        for i in range(parameters_length):
            client_parameters[i].data[:] = initialization_parameters[i].data[:]  # initialization model parameters

    # edge initialization
    edges = []
    cids = np.arange(args.num_clients)
    client_per_edge = int(args.num_clients / args.num_edges)
    p_clients = [0.0] * args.num_edges
    for i in range(args.num_edges):
        selected_client_id = np.random.choice(cids, client_per_edge, replace=False)
        cids = list(set(cids) - set(selected_client_id))
        edges.append(Edge(id=i, client_ids=selected_client_id, total=total, args=args,
                          shared_layers=copy.deepcopy(clients[0].model.shared_layers), global_model=global_model))

        for sid in selected_client_id:
            edges[i].client_register(client=clients[sid])
        edges[i].all_train_sample_num = sum(edges[i].sample_registration.values())
        p_clients[i] = [sample / float(edges[i].all_train_sample_num) for sample in
                        list(edges[i].sample_registration.values())]
        edges[i].refresh_edge()

    # cloud initialization
    cloud = Cloud(shared_layers=copy.deepcopy(clients[0].model.shared_layers),
                  total=total, global_model=global_model, args=args)
    for edge in edges:
        cloud.edge_registration(edge=edge)
    p_edges = [sample / sum(cloud.sample_registration.values()) for sample in
              list(cloud.sample_registration.values())]
    cloud.refresh_cloud()

    train_loss, test_acc, train_time = [0.0] * args.epochs, [0.0] * args.epochs, [0.0] * args.epochs
    complete_time = 0.0

    # save results as excel
    file_dir = "/root/results_waiting_time"
    if not os.path.exists(file_dir):
        os.mkdir(file_dir)
    xlsx_name = "{}_{}_{}_iid[{}]_K1[{}]_K2[{}]_A[{}]_CB[{}]_MC[{}]_MB[{}]_P[{}].xlsx". \
        format(args.dataset, args.model, args.epochs, args.iid, args.num_local_update, args.num_edge_aggregation,
               args.algorithm, args.compute_bias, args.comm_control, args.comm_bias, args.percentage)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Training Results"
    worksheet.cell(row=1, column=1, value='Epoch')
    worksheet.cell(row=1, column=2, value='Train Loss')
    worksheet.cell(row=1, column=3, value='Test Accuracy')
    worksheet.cell(row=1, column=4, value='Completion Time')
    worksheet.cell(row=1, column=5, value='Result k1')
    worksheet.cell(row=1, column=6, value='Result k2')
    worksheet.cell(row=1, column=7, value='Result u')
    worksheet.cell(row=1, column=8, value='Result L')
    worksheet.cell(row=1, column=9, value='Result sigma')
    worksheet.cell(row=1, column=10, value='Average Waiting Time')
    worksheet.cell(row=1, column=11, value='Edge Sync Time')
    worksheet.cell(row=1, column=12, value='Cloud Sync Time')
    worksheet.cell(row=1, column=13, value='Edge Computation Overhead')
    worksheet.cell(row=1, column=14, value='Communication Overhead')
    worksheet.cell(row=1, column=15, value='Edge Waiting Time')
    worksheet.cell(row=1, column=16, value='Cloud Waiting Time')

    # begin training
    for epoch in tqdm(range(args.epochs)):
        cloud.refresh_cloud()
        for edge in edges:
            cloud.edge_registration(edge=edge)

        # client <--> edge
        edge_loss = [0.0] * args.num_edges
        edge_sample = [0] * args.num_edges
        edge_sync_time_temp = [0.0] * args.num_edges
        edge_waiting_time = [0.0] * args.num_edges
        edge_time = [0.0] * args.num_edges
        client_min_time = [0.0] * args.num_edges
        edge_overhead = [0.0] * args.num_edges
        selected_client = []

        for i in range(args.num_edges):
            m = max(int(args.frac * client_per_edge), 1)
            selected_client.append(np.random.choice(edges[i].client_ids, m, replace=False, p=p_clients[i]))
            client_min_time[i] = edges[i].GenerateCompandComm(epoch=epoch)

        min_time = min(client_min_time)

        for i in range(args.num_edges):
            edge_sync_time_temp[i], edge_waiting_time[i] = edges[i].calculate_function(epoch=epoch, min_time=min_time)

        edge_sync_time = sum(edge_sync_time_temp) / (args.num_clients * args.frac)
        cloud.GenerateComm(edges=edges, epoch=epoch)
        cloud_sync_time_sum, k2_list, cloud_waiting_time = cloud.calculation(epoch=epoch)
        for index, edge in enumerate(edges):
            cloud.send_to_edge_k2(edge=edge, index=index)

        for i in range(args.num_edges):
            edge = edges[i]
            edge.refresh_edge()
            client_loss = 0.0

            for idx in selected_client[i]:
                edge.client_register(client=clients[idx])

            while True:
                edge_time[i] += edge.max_time
                edge_overhead[i] += edge.communication_overhead
                temp = 0.0
                for j, idx in enumerate(selected_client[i]):
                    edge.send_to_client(client=clients[idx], index=j)  # send model
                    clients[idx].sync_with_edge()  # update local model
                    loss = clients[idx].local_update(device=device)  # cal
                    temp += loss
                    client_loss += temp / len(selected_client)
                    clients[idx].send_to_edge(edge, epoch=epoch)  # push to edge server
                edge_loss[i] += client_loss
                edge_sample[i] = sum(edge.sample_registration.values())
                flag = edge.aggregate(args)  # aggregation
                if flag:
                    break
            edge_loss[i] /= edges[i].k2
            edge_sync_time_temp[i] *= edges[i].k2

        all_client_loss = sum([e_loss * e_sample for e_loss, e_sample in zip(edge_loss, edge_sample)]) / sum(edge_sample)
        train_loss[epoch] = all_client_loss
        edge_sync_time_sum = sum(edge_sync_time_temp)
        edge_communication_overhead = sum(edge_overhead)
        print(f"The Training Loss is {all_client_loss}")

        # edge <--> cloud
        for edge in edges:
            edge.send_to_cloud(cloud, epoch=epoch)

        k1, u, l, sigma = cloud.aggregate(epoch=epoch, global_loss=all_client_loss)
        complete_time += cloud.max_time
        train_time[epoch] = complete_time
        cloud_sync_time = cloud_sync_time_sum / args.num_edges
        average_waiting_time = (cloud_sync_time_sum + edge_sync_time_sum) / (args.num_clients * args.frac)
        communication_overhead = sum(edge_overhead) + args.num_edges * total * 4
        for index, edge in enumerate(edges):
            cloud.send_to_edge(edge=edge, index=index)

        global_model.load_state_dict(state_dict=copy.deepcopy(cloud.shared_state_dict))
        global_model.eval()
        correct_all, total_all = global_model_test(v_test_loader=v_test_loader, global_model=global_model, device=device)
        avg_acc = correct_all / total_all
        test_acc[epoch] = avg_acc

        row = worksheet.max_row + 1
        worksheet.cell(row=row, column=1, value=epoch)
        worksheet.cell(row=row, column=2, value=all_client_loss)
        worksheet.cell(row=row, column=3, value=avg_acc)
        worksheet.cell(row=row, column=4, value=complete_time)
        worksheet.cell(row=row, column=5, value=k1)
        worksheet.cell(row=row, column=6, value=str(k2_list))
        worksheet.cell(row=row, column=7, value=u)
        worksheet.cell(row=row, column=8, value=l)
        worksheet.cell(row=row, column=9, value=sigma)
        worksheet.cell(row=row, column=10, value=average_waiting_time)
        worksheet.cell(row=row, column=11, value=edge_sync_time)
        worksheet.cell(row=row, column=12, value=cloud_sync_time)
        worksheet.cell(row=row, column=13, value=edge_communication_overhead)
        worksheet.cell(row=row, column=14, value=communication_overhead)
        worksheet.cell(row=1, column=15, value=str(edge_waiting_time))
        worksheet.cell(row=1, column=16, value=cloud_waiting_time)

        path = os.path.join(file_dir, xlsx_name)
        workbook.save(path)

        print("The Global Model Test Acc is {:.2f}%".format(100 * avg_acc))
        print("This global communication round time is {:.4f}s".format(train_time[epoch]))

        # if args.algorithm != 3 and epoch >= 10 and abs(test_acc[epoch] - test_acc[epoch - 10]) <= 0.02 / 100:
        #     print("The model is already convergence!!!")
        #     break

    print("The final Acc is {:.2f}%".format(100 * (max(test_acc))))


def main():
    args = args_parser()
    myFedAvg(args)


if __name__ == '__main__':
    main()
